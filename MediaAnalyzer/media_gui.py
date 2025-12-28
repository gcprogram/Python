import os
import csv
import threading
import queue
import time
from pathlib import Path
import torch
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from tkinter import (
    Tk, Frame, Button, Label, filedialog, ttk, messagebox, Text,
    Scrollbar, Checkbutton, IntVar, StringVar, Menu, Toplevel, Canvas, PhotoImage, END, BOTH, X, Y, RIGHT, BOTTOM, W
)
from tkinter import font as tkfont
import math
from tqdm import tqdm
from PIL import Image, ImageTk, ExifTags
import exiftool  # Bester Allrounder, erfordert separate ExifTool-Installation!

import media_tools
from ai_tools import AITools
from media_tools import _get_exif_data, _get_video_metadata, get_meta_data_bundle, get_kind_of_media, format_time2mmss, \
    extract_mp3_front_cover, read_ai_metadata
from moviepy.video.io.VideoFileClip import VideoFileClip

import logging

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
log = logging.getLogger(__name__)

class MediaAnalyzerGUI:
    def __init__(self, root):
        log.info("Starting Media Analyzer...")
        self.root = root
        self.root.title("🧭 Medien Analyse")
        self.root.geometry("1200x750")
        # Cache für Thumbnail-Pfade
        self._last_thumb_path = None
        self._last_thumb_image = None
        self._model_loading = False
        self.folder:Path = None
        self.create_menu()
        self.create_top_controls()
        self.create_table()
        self._init_styles()
        log.info("AI Models loading...")
        self.aitools = AITools(audio_model_size=self.model_var.get())
        log.info("AI Models loaded...")
        self.current_folder:Path = None
        self.transcripts_missing = 0 # number of audio transcriptions still not processed.
        self.audio_queue = queue.Queue()
        self.audio_workers = []
        self.MAX_AUDIO_WORKERS = 1  # <-- sehr wichtig
        self.start_audio_workers()  # Startet den Thread zur Verarbeitung von Audios/Videos

    #
    # ---------------- Menü ----------------
    #
    def create_menu(self):
        log.info("Menu creating...")
        menubar = Menu(self.root)
        filemenu = Menu(menubar, tearoff=0)
        filemenu.add_command(label="Bulk Load", command=self.choose_folder)
        filemenu.add_command(label="Single File", command=self.choose_single_file)
        filemenu.add_separator()
        filemenu.add_command(label="Exit", command=self.root.quit)
        menubar.add_cascade(label="File", menu=filemenu)
        self.root.config(menu=menubar)
        # Hilfe-Menü
        helpmenu = Menu(menubar, tearoff=0)
        helpmenu.add_command(label="What is this?", command=self.show_help_info)
        helpmenu.add_command(label="Info BLIP", command=self.show_help_blip)
        helpmenu.add_command(label="Info WHISPER", command=self.show_help_whisper)
        helpmenu.add_command(label="Info GPU", command=self.show_help_gpu)
        helpmenu.add_separator()
        helpmenu.add_command(label="Info", command=self.show_info)
        menubar.add_cascade(label="Help", menu=helpmenu)
        self.root.config(menu=menubar)
        self.root.update_idletasks()
        log.info("Menu created.")

    def show_help_info(self):
        messagebox.showinfo("Was ist das?", "Mit dieser Anwendung kannst Du Deine Medienalben\n"
                                                        "Fotos, Videos und sogar Audiodateien\n"
                                                        "automatisch beschreiben lassen. Dazu werden zwei\n"
                                                        "KI-Modelle genutzt, die lokal ausgeführt werden.\n"
                                                        "Deine Daten sind sicher. Ausgabe ist eine exportierbare\n"
                                                        "Liste, die Filename, Orte, Bild- und Audio-Inhalt erkennt.\n"
                                                        "Audios können als Text ausgegeben werden (multilingual).\n"
                                                        "Aus Videos werden Einzelbilder herausgezogen und beschrieben.\n\n"
                                                        "Beispiel Anwendungsfälle:\n"
                                                        "Urlaubsbilder: Ort herausfinden und Kurzbeschreibung, was darauf ist.\n"
                                                        "Lernvideos: Transkipt erstellen und alle 30 sec Bild speichern.\n"
                                                        "Hörbuch: In Text verwandeln.")

    def show_help_blip(self):
        messagebox.showinfo("Info", "Die Offline KI BLIP macht kurze Bildbeschreibungen.\n" 
            "Man kann die maximale Anzahl Tokens anpassen, aber in der Realität verändert sich nicht viel.\n"
            "Voreinstellung hier ist 100.\n"
            "\nBLIP  wurde von Salesforce Research entwickelt und\n"
            "steht unter der Apache License 2.0 https://spdx.org/licenses/Apache-2.0.html\n"
                            )
    def show_help_whisper(self):
        messagebox.showinfo("Info", "Die Offline KI Whisper transkribiert Audios.\n" 
            "️Mögliche Whisper-Modelle zur Auswahl:\n"
            "Name	Größe (RAM)	Geschwindigkeit	Genauigkeit	Bemerkung\n"
            "tiny\t	~75 MB	\t⚡ Sehr schnell	\t😐 Gering	Nur für sehr saubere, kurze Audios ohne Akzent\n"
            "base\t	~142 MB	\t🔸 Schnell	\t🙂 Mittelmäßig	Gute Wahl bei klarer Sprache, ruhiger Umgebung\n"
            "small\t	~466 MB	\t⚖️ Ausgewogen	\t👍 Gut	Sehr gutes Preis-Leistungs-Verhältnis (Standard)\n"
            "medium\t	~1.5 GB	\t🐢 Langsamer	\t💪 Sehr gut	Bessere Erkennung bei Akzenten & Hintergrundgeräuschen\n"
            "large / large-v2 / large-v3	~3–6 GB	\t🐌 Deutlich langsamer	\t🧠 Exzellent	Fast professionelle Qualität, sehr robust bei Rauschen, Akzent, Mehrsprachigkeit\n"
            "\nWhisper wurde von OpenAI entwickelt und\n"
            "          steht unter der MIT Lizenz https://spdx.org/licenses/MIT.html\n"
        )

    def show_help_gpu(self):
        messagebox.showinfo("Info",
            "For using the GPU power of your graphic card for the AI models,\n"
            "you may need to install torch, see below.\n"
            "The models are also usable on a standard CPU.\n"
            "Time for image/frame analysis is 1-2 sec on a Surface Pro 8.\n"
            "Time for audio transcripts of videos is approx. 1.5x the duration.\n"
            "Long videos may not work.\n\n"
            " pip uninstall torch torchvision torchaudio -y\n"
            " pip install torch torchvision torchaudio \n"
            "   --index-url https://download.pytorch.org/whl/cu121\n"
            "Maybe cu118 or cu123 works better on your PC."
        )

    def show_info(self):
        messagebox.showinfo("Info", "Version 1.0\nErstellt von Stefan Markgraf.\n"
                                 "Lizenz: MIT\n\n"
                                 "Benutzte KI-Modelle und Services:\n"
                                 "* BLIP - Bildbeschreibung\n"
                                 "* WHISPER - Audio Transkripte\n"
                                 "* NOMINATIM - Koordinaten zu Adresse (OpenStreetMaps basiert)")


        # ---------------- Layout / Konfiguration ----------------
    def create_top_controls(self):
        # Set default GUI variables
        log.info("Top controls GUI creating...")
        self.model_var = StringVar(value="large-v3")
        self.save_transcript_var = IntVar(value=1)  # standardmäßig aktiviert
        self.interval_var = StringVar(value="30")
        self.save_frames_var = IntVar(value=0)
        self.save_csv_var = IntVar(value=1)
        self.save_xlsx_var = IntVar(value=1)
        self.save_tags_var = IntVar(value=0)

        # Create the config panel
        self.config_frame = Frame(self.root)
        self.config_frame.pack(fill=X, pady=10, padx=10)

        # --- Zeile 1: Transkriptionsmodell ---
        Label(self.config_frame, text="🎙 Transcription Quality:", font=("Arial", 11)).grid(row=0, column=0, sticky=W, padx=5)
        self.model_menu = ttk.Combobox(
            self.config_frame,
            textvariable=self.model_var,
            values=["tiny", "base", "small", "medium", "large-v3"],
            state="readonly", width=10
        )
        self.model_menu.bind("<<ComboboxSelected>>", self.on_whisper_model_change)
        self.model_menu.grid(row=0, column=1, sticky=W, padx=5)
        Checkbutton(self.config_frame, text="Save Transcripts", variable=self.save_transcript_var).grid(row=0, column=2, sticky=W)
        self.create_gpu_status_widget()

        # --- Zeile 2: Video Analyse ---
        Label(self.config_frame, text="⏱ Video Analyse Interval (sec):", font=("Arial", 11)).grid(row=1, column=0, sticky=W, padx=5)
        ttk.Entry(self.config_frame, textvariable=self.interval_var, width=6).grid(row=1, column=1, sticky=W, padx=5)
        Checkbutton(self.config_frame, text="Save Frames", variable=self.save_frames_var).grid(row=1, column=2, sticky=W)

        # --- Zeile 3: Ordner/File Wahl ---
        Label(self.config_frame, text="📂 Analyse File/Ordner:", font=("Arial", 11)).grid(row=2, column=0, sticky=W, padx=5, pady=(10,0))
        Button(self.config_frame, text="Folder select", command=self.choose_folder).grid(row=2, column=1, sticky=W, padx=5, pady=(10,0))
        Button(self.config_frame, text="File select", command=self.choose_single_file).grid(row=2, column=2, sticky=W, padx=5, pady=(10, 0))
        Button(self.config_frame, text="Delete AI tags", command=self.export_treeview_to_delete).grid(row=2, column=4,
                                                                                                      sticky=W, padx=5,
                                                                                                      pady=(10, 0))
        Button(self.config_frame, text="Save AI tags", command=self.export_treeview_to_files).grid(row=2, column=6, sticky=W, padx=5, pady=(10, 0))

        # --- Zeile 4 - Fortschrittanzeige und Ergebnisspeicherung
        self.status_label = Label(self.config_frame, text="Status", font=("Arial", 10))
        self.status_label.grid(row=3, column=0, sticky=W, padx=5, pady=(10, 0))
        self.status_label2 = Label(self.config_frame, text="Fortschritt:", font=("Arial", 10))
        self.status_label2.grid(row=3, column=1, sticky=W, padx=5, pady=(10, 0))
        self.progress = ttk.Progressbar(self.config_frame, orient="horizontal", length=400, mode="determinate")
        self.progress.grid(row=3, column=2, sticky=W, padx=5, pady=(10,0))

        Checkbutton(self.config_frame, text="Save CSV", variable=self.save_csv_var).grid(row=3, column=4, sticky=W)
        Checkbutton(self.config_frame, text="Save Excel", variable=self.save_xlsx_var).grid(row=3, column=5, sticky=W)
        Checkbutton(self.config_frame, text="Save AI Tags (Files)", variable=self.save_tags_var).grid(row=3, column=6, sticky=W)
        self.root.update_idletasks()
        log.info("Top controls GUI created.")

    # ---------------- Tabelle ----------------
    def create_table(self):
        log.info("Table view creating...")
        table_frame = Frame(self.root)
        table_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)

        cols = ("File", "Type", "Date", "Lat", "Lon", "Length", "Address", "Image", "Audio")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=20)
        for col in cols:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_column(c, False))
            width = {"File": 105, "Type": 35, "Date": 100, "Lat":55, "Lon":55, "Length": 50, "Address": 150, "Image": 250, "Audio": 250}.get(col, 100)
            self.tree.column(col, width=width, anchor="w")

        self.setup_click_play()
        # Scrollbars anlegen
        vsb = Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)

        # Treeview <-> Scrollbars koppeln
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        # Grid-Layout: tree in (0,0), vsb in (0,1), hsb in (1,0)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        # Thumbnail-Hover
        self.thumb_window = None
        self._last_thumb_path = None
        self._last_thumb_image = None
        self.tree.bind("<Motion>", self.on_hover)
        self.tree.bind("<Leave>", self.on_leave)

        # Double-Click Play
        self.tree.bind("<Double-1>", self.on_double_click)

        # Make the tree expand when window is resized
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)
        self.root.update_idletasks()
        log.info("Table view created.")
    #
    # Reads data from TreeView.
    #
    def get_treeview_data(self, tree: ttk.Treeview):
        columns = tree["columns"]
        headers = [tree.heading(col)["text"] or col for col in columns]

        rows = []
        for item_id in tree.get_children():
            rows.append(tree.item(item_id, "values"))

        return headers, rows

    #
    # Writes CSV from filled Treeview.
    #
    def export_treeview_to_csv(self, file_path: str):
        headers, rows = self.get_treeview_data(self.tree)

        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(headers)
            writer.writerows(rows)

    #
    # Writes Excel from filled Treeview.
    #
    def export_treeview_to_xlsx(self, file_path: str):
        headers, rows = self.get_treeview_data(self.tree)

        wb = Workbook()
        ws = wb.active
        ws.title = "Media Analysis"

        # Header
        ws.append(headers)

        # Daten
        for row in rows:
            ws.append(list(row))

        # Spaltenbreite automatisch anpassen
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row in rows:
                try:
                    max_len = max(max_len, len(str(row[col_idx - 1])))
                except IndexError:
                    pass

            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        wb.save(file_path)

    #
    # Writes AI tags for Address, Image and Audio to image, video, audio files
    #
    def export_treeview_to_files(self):
        headers, rows = self.get_treeview_data(self.tree)

         # Daten
        with exiftool.ExifTool() as et:
            for row in rows:
                a = list(row)
                address:str = a[6]
                image_text = a[7]
                audio_text = a[8]
                media_tools.write_ai_metadata(os.path.join(self.folder, a[0]), address,image_text,audio_text,et )

    #
    # Delete AI tags from shown Images, Videos and Audio in TreeView
    #
    def export_treeview_to_delete(self):
        headers, rows = self.get_treeview_data(self.tree)

         # Daten
        with exiftool.ExifToolHelper() as et:
            for row in rows:
                a = list(row)
                media_tools.delete_ai_metadata(os.path.join(self.folder, a[0]), et)

    def get_gpu_status(self):
        if torch.cuda.is_available():
            name = torch.cuda.get_device_name(0)
            return True, name
        return False, None

    def create_gpu_status_widget(self):
        gpu_available, gpu_name = self.get_gpu_status()

        if gpu_available:
            text = f"GPU: {gpu_name} | CUDA aktiv"
            style = "GpuActive.TLabel"
        else:
            text = "GPU: nicht verfügbar – CPU-Modus"
            style = "GpuInactive.TLabel"

        label = ttk.Label(self.config_frame, text=text, style=style)
        label.grid(row=0, column=2, sticky="e", padx=5, pady=(10, 0))
        self.gpu_status_label = label

    def update_gpu_status_label(self):
        if not hasattr(self, "aitools"):
            return

        device = getattr(self.aitools.audio_model, "device", None)

        if device and device.type == "cuda":
            text = "GPU: CUDA aktiv"
            style = "GpuActive.TLabel"
        else:
            text = "GPU: CPU-Modus"
            style = "GpuInactive.TLabel"

        self.gpu_status_label.config(text=text, style=style)

    def _init_styles(self):
        style = ttk.Style()

        style.configure(
            "GpuActive.TLabel",
            background="darkgreen",
            foreground="black",
            font=("Segoe UI", 9, "bold"),
        )

        style.configure(
            "GpuInactive.TLabel",
            background="red",
            foreground="white",
            font=("Segoe UI", 9),
        )

    def on_whisper_model_change(self, event=None):
        """
        Wird aufgerufen, wenn ein anderes Whisper-Modell ausgewählt wird.
        Lädt das Modell im Hintergrund.
        """
        if self._model_loading:
            return  # Mehrfachklicks ignorieren

        self._model_loading = True
        whisper_choice: str = self.model_var.get()

        self.status_label.config(
            text=f"🎙 Lade Audio-Modell '{whisper_choice}' …"
        )
        self.progress.config(mode="indeterminate")
        self.progress.start(10)
        self.root.update_idletasks()

        def _load_model():
            try:
                self.aitools._load_audio_model(audio_model_size=whisper_choice)
                #self.aitools = AITools(audio_model_size=whisper_choice)
                # GUI-Update: GPU-Status
                self.root.after(0, self.update_gpu_status_label)

            except Exception as e:
                self.root.after(
                    0,
                    lambda e=e: messagebox.showerror(
                        "ERROR",
                        f"Audio Modell konnte nicht geladen werden:\n{e}"
                    )
                )
            finally:
                self.root.after(0, self._on_whisper_model_loaded)

        threading.Thread(
            target=_load_model,
            daemon=True,
            name="WhisperModelSwitch"
        ).start()

    def _on_whisper_model_loaded(self):
        self.progress.stop()
        self.progress.config(mode="determinate")
        self.progress["value"] = 0

        model_name = self.model_var.get()
        self.status_label.config(
            text=f"✅ Whisper '{model_name}' bereit"
        )

        self._model_loading = False

    # ---- Thumbnail Hover ----
    def on_hover(self, event):
        """
        Erweitertes Hover-Verhalten:
        - #1 (Datei): Thumbnail (Bild oder Video-Mittel-Frame)
        - #8 (Beschreibung) oder #9 (Transkript): Text-Tooltip mit kompletter Beschreibung
        Debounce: gleiche Datei/Text nicht ständig neu laden.
        """
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            self.hide_thumbnail()
            self.hide_text_tooltip()
            return

        row_id = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not row_id or not col:
            self.hide_thumbnail()
            self.hide_text_tooltip()
            return

        values = self.tree.item(row_id, "values")
        if not values:
            self.hide_thumbnail()
            self.hide_text_tooltip()
            return

        # Relativer Pfad/Dateiname ist in Spalte 0 (Datei)
        filename = values[0]
        # Die Textfelder sind in Spalte 7 (Beschreibung) und 8 (Transkript) - tree columns sind 1-indexed (#1..)
        desc_col = "#8"
        text_col = "#9"

        # Wenn über Dateispalte -> Thumbnail anzeigen
        if col == "#1":
            # Verstecke Text-Tooltip, falls sichtbar
            self.hide_text_tooltip()

            folder = getattr(self, "current_folder", getattr(self, "folder", ""))
            path = os.path.join(folder, filename)
            if not os.path.exists(path):
                self.hide_thumbnail()
                return

            # Debounce für Thumbnail: identischer Pfad -> nichts tun
            if self._last_thumb_path == path:
                return
            self._last_thumb_path = path

            # Image vs. Video thumbnail
            if filename.lower().endswith((".jpg", ".jpeg", ".png")):
                self.show_image_thumbnail(path, event.x_root, event.y_root)
            elif filename.lower().endswith((".mp4", ".mov", ".avi")):
                self.show_video_thumbnail(path, event.x_root, event.y_root)
            elif filename.lower().endswith((".mp3", ".wav", ".m4a", ".flac")):
                cover = extract_mp3_front_cover(path)
                if cover:
                    self.show_image_thumbnail(cover, event.x_root, event.y_root)
                else:
                    self.hide_thumbnail()
            else:
                self.hide_thumbnail()
            return

        # Wenn über Beschreibung oder Transkript -> Text-Tooltip anzeigen
        if col in (desc_col, text_col):
            # Verstecke Thumbnail, falls sichtbar
            self.hide_thumbnail()

            # Den kompletten Text aus der jeweiligen Spalte holen
            # tree.set benötigt den item id und die column name (hier nutzen wir index via values)
            # values ist ein tuple entsprechend der Spaltenreihenfolge
            # Beschreibung = values[7], Transkript = values[8]
            try:
                if col == desc_col:
                    full_text = values[7] or ""
                else:
                    full_text = values[8] or ""
            except Exception:
                full_text = ""

            # Wenn kein Text -> nichts anzeigen
            if not full_text.strip():
                self.hide_text_tooltip()
                return

            # Debounce: gleiches Text-Tooltip vermeiden
            if getattr(self, "_last_tooltip_text", None) == full_text:
                return
            self._last_tooltip_text = full_text

            # Zeige Tooltip (mit max Größe und Scrollbar falls nötig)
            self.show_text_tooltip(full_text, event.x_root + 15, event.y_root + 15)
            return

        # In allen anderen Fällen beides verbergen
        self.hide_thumbnail()
        self.hide_text_tooltip()

    ###########################################
    # Zeigt Tooltips von Image und Audio Description.
    ###########################################
    def show_text_tooltip(self, text:str, x:int, y:int, max_width_px:int=600, max_height_px:int=400):
        if not text.strip():
            return

        self.hide_text_tooltip()

        win = Toplevel(self.root)
        win.wm_overrideredirect(True)
        win.attributes("-topmost", True)

        frame = Frame(win, bd=1, relief="solid", bg="white")
        frame.pack(fill="both", expand=True)

        vsb = Scrollbar(frame, orient="vertical")
        vsb.pack(side="right", fill="y")

        ta = Text(
            frame,
            wrap="word",
            yscrollcommand=vsb.set,
            padx=8,
            pady=6,
            bd=0,
            width=1,  # WICHTIG: minimale Startbreite
            height=1,
            bg="white"
        )
        ta.pack(side="left", fill="both", expand=True)
        vsb.config(command=ta.yview)

        ta.insert("1.0", text)
        ta.config(state="disabled")

        self.text_tooltip_window = win
        self.text_tooltip_window_text = ta

        # Jetzt erst Layout berechnen lassen
        win.update_idletasks()

        self._adjust_tooltip_size(max_width_px, max_height_px)

        win.geometry(f"+{x}+{y}")

    ###########################################
    # Berechnet die Größe des Tooltips
    ###########################################
    def _adjust_tooltip_size(self, max_width_px:int, max_height_px:int):
        win = self.text_tooltip_window
        ta = self.text_tooltip_window_text

        font = tkfont.Font(font=ta.cget("font"))
        line_height = font.metrics("linespace")

        # Maximale Breite setzen → erzwingt Wortumbruch
        ta.config(width=1)
        win.geometry(f"{max_width_px}x{max_height_px}")
        win.update_idletasks()

        # Zeilenanzahl ermitteln
        last_index = ta.index("end-1c")
        total_lines = int(last_index.split(".")[0])

        needed_height = total_lines * line_height + 20
        final_height = min(needed_height, max_height_px)

        win.geometry(f"{max_width_px}x{final_height}")

    def hide_text_tooltip(self):
        """Versteckt den Text-Tooltip falls vorhanden."""
        if getattr(self, "text_tooltip_window", None):
            try:
                self.text_tooltip_window.destroy()
            except:
                pass
            self.text_tooltip_window = None
        self._last_tooltip_text = None

    def show_video_thumbnail(self, path, x:int, y:int):
        try:
            clip = VideoFileClip(path)
            mid = clip.duration / 2
            frame = clip.get_frame(mid)  # numpy array
            clip.close()
            img = Image.fromarray(frame)
            img.thumbnail((200, 200))
            photo = ImageTk.PhotoImage(img)
            self._last_thumb_image = photo
            self._show_thumbnail_window(photo, x, y)
        except Exception:
            self.hide_thumbnail()

    def _show_thumbnail_window(self, photo, x:int, y:int):
        if self.thumb_window:
            self.thumb_window.destroy()
        self.thumb_window = Toplevel(self.root)
        self.thumb_window.overrideredirect(True)
        self.thumb_window.geometry(f"+{x + 20}+{y + 20}")
        label = Label(self.thumb_window, image=photo)
        label.image = photo
        label.pack()

    def on_leave(self, event):
        self.hide_thumbnail()
        self._last_thumb_path = None
        self._last_thumb_image = None

    def show_image_thumbnail(self, path, x:int, y:int):
        try:
            img = Image.open(path)
            # Orientierung aus EXIF
            try:
                for orientation in ExifTags.TAGS.keys():
                    if ExifTags.TAGS[orientation] == 'Orientation':
                        break
                exif = img._getexif()
                if exif and orientation in exif:
                    o = exif[orientation]
                    if o == 3:
                        img = img.rotate(180, expand=True)
                    elif o == 6:
                        img = img.rotate(270, expand=True)
                    elif o == 8:
                        img = img.rotate(90, expand=True)
            except Exception:
                pass
            img.thumbnail((200, 200))
            photo = ImageTk.PhotoImage(img)
            self._last_thumb_image = photo
            self._show_thumbnail_window(photo, x, y)
        except Exception:
            self.hide_thumbnail()

    def hide_thumbnail(self):
        if self.thumb_window:
            self.thumb_window.destroy()
            self.thumb_window = None

    def setup_click_play(self):
        self.tree.bind("<Double-1>", self.on_double_click)

    def on_double_click(self, event):
        row_id = self.tree.identify_row(event.y)
        if not row_id:
            return
        values = self.tree.item(row_id, "values")
        filename = values[0]
        folder = self.folder
        path = os.path.join(folder, filename)
        if os.path.exists(path) and filename.lower().endswith((".mp4", ".mov", ".avi")):
            # Standard Video-Player öffnen
            if os.name == "nt":  # Windows
                os.startfile(path)
            elif os.name == "posix":  # macOS/Linux
                print("masOS, Linux Playback not implemented yet")
                #subprocess.run(["open" if sys.platform == "darwin" else "xdg-open", path])

    def set_process(self, value):
        self.progress["value"] = value

    # ---- Tabellen-Sortierung ----
    def sort_column(self, col, reverse):
        items = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
        try:
            items.sort(key=lambda t: (float(t[0]) if t[0].replace('.', '', 1).isdigit() else t[0].lower()),
                        reverse=reverse)
        except Exception:
            items.sort(key=lambda t: t[0].lower(), reverse=reverse)
        for index, (val, k) in enumerate(items):
            self.tree.move(k, '', index)
        self.tree.heading(col, command=lambda: self.sort_column(col, not reverse))

    # ---------------- Analyse ----------------
    def choose_folder(self):
        self.folder = filedialog.askdirectory(title="Verzeichnis wählen")
        if self.folder:
            self.current_folder = self.folder
            threading.Thread(target=self.analyze_folder, args=(self.folder,), daemon=True).start()

    def choose_single_file(self):
        file = filedialog.askopenfilename(
            title="Datei wählen",
            filetypes=[("Medien", "*.jpg *.jpeg *.png *.mp4 *.mov *.avi *.mp3 *.wav *.m4a *.flac"), ("Alle Dateien", "*.*")]
        )
        if file:
            self.folder = os.path.dirname(file)
            self.current_folder = self.folder
            threading.Thread(target=self.analyze_single_file, args=(file,), daemon=True).start()

    #
    # Hauptroutine liest files vom file_path und füllt die Tabelle.
    #
    def analyze_folder(self, file_path:Path):
        self.status_label.config(text="📦 Lade Files...")
        self.root.update_idletasks()
        self.tree.delete(*self.tree.get_children())

        if self._model_loading:
            messagebox.showwarning(
                "Bitte warten",
                "Das Audio-Modell wird noch geladen."
            )
            return

        interval = int(self.interval_var.get())

        if os.path.isdir(file_path):
            # Erzeuge eine Liste von Dateipfaden für alle Dateien im Verzeichnis
            folder = file_path
            all_files = [os.path.join(root, f) for root, _, files in os.walk(file_path) for f in files]
        elif os.path.isfile(file_path):
            # Setze all_files auf eine Liste, die nur den angegebenen Dateipfad enthält
            folder = os.path.dirname(file_path)
            all_files = [file_path]
        else:
            # Bei ungültigem Pfad kann hier ein Fehler ausgelöst oder behandelt werden
            all_files = []
            print("Der angegebene Pfad ist weder eine Datei noch ein Verzeichnis.")
            return

        total = len(all_files)
        self.progress["maximum"] = total
        self.progress["value"] = 0
        self.root.update_idletasks()

        self.status_label.config(text=f"🔍 Analysiere {total} Dateien...")

        with exiftool.ExifTool() as et:

            self.transcripts_missing = 0
            transcripts_cnt = 0
            transcripts_duration = 0
            for i, path in enumerate(tqdm(all_files, desc="Analysiere")):
                # Startzeit
                start_time = time.time()
                kind = get_kind_of_media(path)
                if kind == "unknown":
                    self.progress["value"] = i + 1
                    continue
                relpath = os.path.relpath(path, folder)
                rec = {"File": relpath, "Type": kind.capitalize(), "Date": "", "Lat": "", "Lon": "", "Length": "", "Address": "", "Image": "", "Audio": ""}
                item_id = self.tree.insert("", "end", values=tuple(rec.values()))
                image_text = ""
                audio_text = ""
                try:
                    meta = get_meta_data_bundle(path, et_instance=et)
                    rec["Date"] = meta.get("Date", "")
                    rec["Address"] = meta.get("Address", "")
                    rec["Lat"] = meta.get("Lat", "")
                    rec["Lon"] = meta.get("Lon", "")
                    rec["Length"] = meta.get("Length", "")

                    self._update_tree_columns(item_id, rec)

                    meta_ai = read_ai_metadata(path, et)
                    address:str = meta_ai.get("address", "")
                    image_text:str = meta_ai.get("caption", "")
                    audio_text:str = meta_ai.get("transcript", "")
                    if len(image_text) < 4:
                        # Mache Image Beschreibung sofort
                        if kind == "image":
                            image_text = self.aitools.describe_image(path)
                        elif kind == "video":
                            # Das ist aufwendiger: Video zerlegen in Einzelbilder, alle %interval%s Sekunden.
                            image_text = self.aitools.describe_video_by_frames(path, interval)
                            if self.save_frames_var.get():
                                self._save_video_frames(path, interval)
                        elif kind == "audio" and len(audio_text) < 4:
                            print("Versuche Bild aus Audio zu extrahieren")
                            image = extract_mp3_front_cover(path)
                            if image is None:
                                log.warning(f"{relpath} has no image")
                                image_text = ""
                            else:
                                image_text = self.aitools.describe_image(image)
                                log.info(f"Cover-Bild zeigt: {image_text}")
                                # MP3 Cover Image extrahieren und beschreiben.

                    rec["Image"] = image_text
                    rec["Audio"] = audio_text
                    self._update_tree_columns(item_id, rec)

                    # Nur audio_text im Hintergrund erzeugen
                    if kind in ("video", "audio"):
                        if audio_text == "..." or audio_text == "":
                            self.transcripts_missing += 1
                            transcripts_cnt += 1
                            transcripts_duration:int = transcripts_duration+ int(rec["Length"])
                            self._run_ai_analysis(path, kind, item_id, image_text, float(rec["Length"]))


                except Exception as e:
                    print("⚠️ Fehler bei:", path, e)

                self.progress["value"] = i + 1
                self.root.update_idletasks()

        self.status_label.config(
            text=f"🗣️ Transcribe Audios {self.transcripts_missing}"
        )
        self.progress["maximum"] = transcripts_duration
        self.progress["value"] = 1
        #self.progress.config(mode="indeterminate")

        # --- Warten, bis alle Audio-Jobs abgearbeitet sind ---
        def _wait_for_audio_jobs():
            self.audio_queue.join()  # ⬅ BLOCKIERT, bis alle task_done() aufgerufen wurden
            self.root.after(0, _on_all_audio_done)

        def _on_all_audio_done():
            if self.save_csv_var.get():
                out_path = os.path.join(folder, "_media_analysis.csv")
                self.export_treeview_to_csv(out_path)
                self.status_label.config(text=f"✅ Fertig → {os.path.basename(out_path)}")
            if self.save_xlsx_var.get():
                out_path = os.path.join(folder, "_media_analysis.xlsx")
                self.export_treeview_to_xlsx(out_path)
                self.status_label.config(text=f"✅ Fertig → {os.path.basename(out_path)}")
            if self.save_tags_var.get():
                self.export_treeview_to_files()
                self.status_label.config(text=f"✅ Fertig → Tags in Files")

            messagebox.showinfo(
                "Fertig",
                "✅ Analyse abgeschlossen.\nAlle Audio-Transkripte sind fertig."
            )

        threading.Thread(
            target=_wait_for_audio_jobs,
            daemon=True,
            name="WaitForAudioJobs"
        ).start()

        # HIER WARTEN und progress updaten, während die Audio Transcripts erzeugt werden im Thread _audio_worker_loop().
        self.progress.config(mode="determinate")
        if total == 1:  # only show this window in single file mode
            text = audio_text + "\n\n " + image_text.translate(str.maketrans("|", '\n'))
            self.show_result_window(file_path, kind, text)

    # ---------------- Single File Mode ----------------
    def analyze_single_file(self, file_path:Path):
        self.analyze_folder(file_path)

    #
    # Startet einen Thread um Audio Transkription zu machen
    #
    def start_audio_workers(self):
        for _ in range(self.MAX_AUDIO_WORKERS):
            t = threading.Thread(
                target=self._audio_worker_loop,
                daemon=True,
                name="AudioWorker"
            )
            t.start()
            self.audio_workers.append(t)

    #
    # THREAD: Macht die eigentliche Transkription und holt Jobs aus der Queue
    #
    def _audio_worker_loop(self):
        while True:
            job = self.audio_queue.get()
            if job is None:
                print("_audio_worker_loop(): No audio work to do")
                break  # sauberer Shutdown

            path, kind, item_id, image_text, duration = job
            audio_text = ""
            try:

                audio_text = self.aitools.transcribe_audio(path)
                print(f"_audio_worker_loop(): {os.path.basename(path)}\n{audio_text}")
                if self.save_transcript_var.get():
                    text = f"\"{audio_text}\"\n\n{image_text.translate(str.maketrans('|', '\n'))}"
                    with open(f"{path}.txt", "w", encoding="utf-8") as f:
                        f.write(text)

            except Exception as e:
                audio_text = f"⚠️ Fehler: {e}"
            finally:
                self.root.after(
                    0,
                    self._update_tree_ai_columns,
                    item_id,
                    audio_text
                )
                self.transcripts_missing -= 1  # ✅ hierhin
                self.progress["value"] += int(duration)
                self.audio_queue.task_done()

    def _update_tree_ai_columns(self, item_id, audio_text):
        values = list(self.tree.item(item_id, "values"))
        values[8] = audio_text
        self.tree.item(item_id, values=values)

        # Fortschritt
        self.progress["value"] += 1

        if self.progress["value"] >= self.progress["maximum"]:
            self.status_label.config(text="✅ Audio-Transkription abgeschlossen")

    #
    # Push jedes Audio und Video in die Queue. _audio_worker_loop()
    #
    def _run_ai_analysis(self, path, kind:str, item_id, image_text:str, length:float):
        if kind not in ("audio", "video"):
            return
        self.audio_queue.put(
            (path, kind, item_id, image_text, length)
        )
        log.debug(f"queue.size={self.audio_queue.qsize()}")

    #
    # AI Ergebnisse eintragen.
    #
    def _update_tree_ai_columns(self, item_id, audio_text:str):
        values = list(self.tree.item(item_id, "values"))

        # Spaltenindex:
        # 0 File | 1 Type | 2 Date | 3 Lat | 4 Lon | 5 Length | 6 Address | 7 Image | 8 Audio
        if audio_text:
            values[8] = audio_text

        self.tree.item(item_id, values=values)

    #
    # AI Ergebnisse eintragen.
    #
    def _update_tree_columns(self, item_id, rec):
        values = list(self.tree.item(item_id, "values"))
        # 0 File | 1 Type | 2 Date | 3 Lat | 4 Lon | 5 Length | 6 Address | 7 Image | 8 Audio
        #rec = {"File": relpath, "Type": kind.capitalize(), "Date": "", "Lat": "", "Lon": "", "Length": "","Address": "", "Image": "", "Audio": ""}

        values[0] = rec["File"]
        values[1] = rec["Type"]
        values[2] = rec["Date"]
        values[3] = rec["Lat"]
        values[4] = rec["Lon"]
        if rec["Length"]:
            minutes:int = int(float(rec["Length"]) // 60)
            seconds:int = int(float(rec["Length"]) % 60)
            values[5] = f"{minutes}:{seconds:02d}"
        else:
            values[5] = ""
        values[6] = rec["Address"]
        values[7] = rec["Image"]
        self.tree.item(item_id, values=values)


    def show_result_window(self, file_path, kind, result_text):
        win = Toplevel(self.root)
        win.title(f"Analyse: {os.path.basename(file_path)}")
        win.geometry("800x600")

        Label(win, text=f"Type: {kind.capitalize()}", font=("Arial", 11, "bold")).pack(pady=5)
        Label(win, text=f"File: {file_path}", font=("Arial", 9)).pack(pady=2)

        text_area = Text(win, wrap="word")
        text_area.pack(fill=BOTH, expand=True, padx=10, pady=10)
        text_area.insert(END, result_text)
        text_area.config(state="disabled")

    # ---------------- Hilfsfunktionen ----------------
    def _save_video_frames(self, video_path, interval):
        """Speichert Frames als PNGs im gleichen Ordner."""
        from moviepy.video.io.VideoFileClip import VideoFileClip
        clip = VideoFileClip(video_path)
        base, _ = os.path.splitext(video_path)
        t = 0
        while t < clip.duration:
            frame = clip.get_frame(t)
            img = Image.fromarray(frame)
            seq = int(t / interval) + 1
            mmss = format_time2mmss(t).replace(":", "-")
            out_name = f"{base}+{mmss}.png"
            img.save(out_name)
            t += interval
        clip.close()

    def _save_transcript(self, file_path, text):
        """Speichert Transkript in .txt-Datei."""
        base, _ = os.path.splitext(file_path)
        txt_path = f"{base}_transkript.txt"
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(text)


if __name__ == "__main__":
    root = Tk()
    app = MediaAnalyzerGUI(root)
    root.mainloop()
